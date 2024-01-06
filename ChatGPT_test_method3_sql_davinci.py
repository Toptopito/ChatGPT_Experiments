import openai
import pandas as pd
import sqlite3
import openpyxl
import time

openai.api_key = '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'

model = 'text-davinci-003'

# create SQLite DB (create a new DB file if it doesn't exist)
conn = sqlite3.connect("antidepressant_chat.db")
cursor = conn.cursor()

# create a table
cursor.execute("""
    CREATE TABLE IF NOT EXISTS antidepressant_chat (
        idx TEXT,
        age_low INTEGER,
        age_high INTEGER,
        age_range TEXT,
        gender TEXT,
        antidepressant TEXT
    )
""")

# Commit the transaction and close the connection
conn.commit()
conn.close()

validation_file_path = "C:/Users/vladc/OneDrive/Documents/GMU Research/AI Chat for Depression/Reference Files/test_validation_file.csv"

validation_df = pd.read_csv(validation_file_path)

# Connec to the SQLite DB
conn = sqlite3.connect("antidepressant_chat.db")

# Insert dataframe into the DB
validation_df.to_sql("antidepressant_chat", conn, if_exists='replace', index=False)

# close the DB connection
conn.close()

# Connec to the SQLite DB
conn = sqlite3.connect("antidepressant_chat.db")
cursor = conn.cursor()


# function to get table columns from SQLite DB
def get_table_columns(table_name):
    cursor.execute("PRAGMA table_info({})".format(table_name))
    columns = cursor.fetchall()
    return[column[1] for column in columns]
    
    
# function to generate SQL query from input text using ChatGPT
def generate_sql_query(table_name, text, columns):
    prompt = """You are a ChatGPT language model that can generate SQL queries.
    You will generate an SQL query from natural language input text to get the relevant antidepressant. 
    Categorize the gender from the input text as either male or female before generating the query.
    The table name is {}.\n
    The column names are {}.\n
    Input text: {}\n
    SQL Query:
    """.format(table_name, columns, text)
    
    incomplete = True
    tries = 0
    while(incomplete):
        try:
            request = openai.Completion.create(
                prompt=prompt,
                temperature=0,
                max_tokens=300,
                top_p=1,
                frequency_penalty=0,
                presence_penalty=0,
                stop=None,
                model=model,
                )           
               
        except Exception as e:
            tries += 1
            print(f"For input {text}, failed. Try #: {tries}")
            print(e)
            raise e

            if tries == 5:
                incomplete = False
                return None
            else:
                time.sleep(5) # wait 5 seconds then try again
                incomplete = True
    
    sql_query = request.choices[0].text.strip()
    
    return sql_query
    

# function to execute SQL query on SQLite DB
def execute_sql_query(query):
    cursor.execute(query)
    result = cursor.fetchall()
    return result
    
table_name = 'antidepressant_chat'
columns = get_table_columns(table_name)

#workbook_path = "C:/Users/vladc/OneDrive/Documents/GMU Research/AI Chat for Depression/AI Chat for Depression Tracking Sheet.xlsx"
#sheet_name = "Test Cases"
#method3_result_col = 10
#start_row = 2
#prompt_col = 'B'

#workbook = openpyxl.load_workbook(workbook_path)
#sheet = workbook[sheet_name]

## loop through sheet and fill the results column
#for row in range(start_row, sheet.max_row+1):

#    test_case_cell = "{}{}".format(prompt_col, row)
#    test_case = sheet[test_case_cell].value

test_case = ""
while(test_case.lower() != 'exit'):
    
    test_case = input("Tell me about your name and gender: ")
    
    sql_query = generate_sql_query(table_name, test_case, columns)
       
    if sql_query:
        result = execute_sql_query(sql_query)
        print("Response: ", result)
    else:
        result = None
    
    print(f"Response: {result}")
    
#    method3_cell = sheet.cell(row, method3_result_col)
#    method3_cell.value = result
        
# close the DB connection
cursor.close()
conn.close()
