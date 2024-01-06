import openai
import os
import pandas as pd
import openpyxl
import time

openai.api_key = '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'

model = 'gpt-3.5-turbo-0613'

validation_file_path = "C:/Users/vladc/OneDrive/Documents/GMU Research/AI Chat for Depression/Reference Files/test_validation_file.csv"

validation_df = pd.read_csv(validation_file_path)

workbook_path = "C:/Users/vladc/OneDrive/Documents/GMU Research/AI Chat for Depression/AI Chat for Depression Tracking Sheet.xlsx"
sheet_name = "Test Cases"
method1_result_col = 6
start_row = 2
prompt_col = 'B'

workbook = openpyxl.load_workbook(workbook_path)
sheet = workbook[sheet_name]

# loop through sheet and fill the results column
for row in range(start_row, sheet.max_row+1):

    test_case_cell = "{}{}".format(prompt_col, row)
    test_case = sheet[test_case_cell].value
    
    prompt = """Please look into the following data:\n {}. 
    Please also look at the following information provided by the user: {}.
    Categorize the age of the user as an age group. Categorize the gender of the user as male or female. 
    Then, based on age group and gender find the row in the data where the gender and age group match the information provided by the user.
    From that row, find the recommended antidepressant that matched the user's age group and gender.
    
    Examples for finding the matching antidepressant:
    Example 1: "age": 39, "gender": male
    Category: "age_group": 20-40, "gender": male 
    Result: "antidepressant": Venlafaxine | Desvenlafaxine

    Example 2: "age": 50, "gender": female
    Category: "age_group": 41-64, "gender": female
    Result: "antidepressant": Sertraline | Desvenlafaxine
    
    Example 3: "age": 87, "gender": female
    Category: "age_group": 80-89, "gender": female
    Result: "antidepressant": Desvenlafaxine
    
    Do not show your your process to the user. Show only the antidepressant.
    "antidepressant":
    """.format(validation_df, test_case)
    
    incomplete = True
    tries = 0
    while(incomplete):
        try:
            request = openai.ChatCompletion.create(
                model = model,
                messages = [
                    {"role": "user", "content": prompt}
                ]
            )
            incomplete = False
        except:
            tries += 1
            print(f"For input {test_case}, failed. Try: {tries}")
            if tries == 5:
                incomplete = False
            else:
                time.sleep(5) # wait 5 seconds then try again
                incomplete = True
        
    if tries < 5:
        response = request['choices'][0]['message']['content']
    else:
        response = None
    
    method1_cell = sheet.cell(row, method1_result_col)
    method1_cell.value = response

workbook.save(workbook_path)
