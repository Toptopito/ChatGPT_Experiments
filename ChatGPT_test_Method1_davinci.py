import openai
import os
import pandas as pd
import openpyxl

openai.api_key = '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'

model = 'text-davinci-003'

validation_file_path = "./data/test_validation_file.csv"

validation_df = pd.read_csv(validation_file_path)

workbook_path = "./data/AI Chat for Depression Tracking Sheet.xlsx"
sheet_name = "Test Cases"
method1_result_col = 6
start_row = 2
prompt_col = 'B'

workbook = openpyxl.load_workbook(workbook_path)
sheet = workbook[sheet_name]

# loop through sheet and fill the results column
for row in range(start_row, sheet.max_row+1):

    test_case_cell = "{}{}".format(prompt_col, row)
    test_case = sheet[test_case_cell].value
    
    prompt = """Please look into the following data:\n {}. 
    Please also look at the following information provided by the user: {}.
    Categorize the age and gender of the user. Gender can only be male or female.
    Then, based on age group and gender find the row in the data where the gender and age group match the information provided by the user.
    From that row, find the recommended antidepressant that matched the user's age group and gender.
    
    Examples for finding the matching antidepressant:
    Example 1: "age": 39, "gender": male
    Result: "age_group": 20-40, "gender": male, "antidepressant": Venlafaxine | Desvenlafaxine

    Example 2: "age": 50, "gender": female
    Result: "age_group": 41-64, "gender": female, "antidepressant": Sertraline | Desvenlafaxine
    
    Example 3: "age": 87, "gender": female
    Result: "age_group": 80-89, "gender": female, "antidepressant": Desvenlafaxine
      
    "antidepressant":
    """.format(validation_df, test_case)
    
    request = openai.Completion.create(
        prompt=prompt,
        temperature=0,
        max_tokens=300,
        top_p=1,
        frequency_penalty=0,
        presence_penalty=0,
        stop=None,
        model=model,
        )
    
    response = request.choices[0].text.strip()
    
    method1_cell = sheet.cell(row, method1_result_col)
    method1_cell.value = response

workbook.save(workbook_path)
